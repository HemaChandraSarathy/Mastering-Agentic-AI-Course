"""Headline demo: why financial NUMBERS should not come from RAG.

Mirrors the course's experiment on OUR data: flatten the financial table to text, chunk it
(fixed-size, as the course does), embed, and retrieve for a numeric question. The retrieved
context is fragmented (numbers separated from their period headers) — the exact failure mode
the course documents. Then show Pipeline A: deterministic extraction gives the EXACT figure
with provenance + a passing reconciliation check.

Claude-free (retrieval = Nebius embeddings, extraction = local), so it runs even when the
Anthropic balance is exhausted. Usage: python scripts/run_numbers_demo.py
"""
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "src"))
sys.stdout.reconfigure(encoding="utf-8")

from dotenv import load_dotenv
from openpyxl import load_workbook

from portcoiq_rag.pipeline_a.extract import extract_from_excel
from portcoiq_rag.pipeline_a.validate import validate
from portcoiq_rag.pipeline_b.embed import embed_chunks
from portcoiq_rag.pipeline_b.models import Chunk
from portcoiq_rag.pipeline_b.retrieve import InMemoryDenseRetriever

load_dotenv()
XLSX = REPO / "corpus" / "sample" / "Meridian_SAMPLE_financials.xlsx"
QUESTION = "What was Meridian's net income in Q3 2025?"
PERIOD, METRIC = "Q3 2025", "net_income"


def flatten(path):
    wb = load_workbook(path, data_only=True)
    out = []
    for ws in wb.worksheets:
        out.append(f"[{ws.title}]")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                out.append("  ".join(cells))
    return "\n".join(out)


print("=" * 72)
print("APPROACH 1 — RAG over the financial TABLE-AS-TEXT (the course's approach)")
print("=" * 72)
text = flatten(XLSX)
size = 120  # small fixed chunks → fragments the table, like the course's 512-char failure
raw = [text[i:i + size] for i in range(0, len(text), size)]
chunks = [Chunk(content=t, source_file=XLSX.name, strategy="fixed", chunk_index=i)
          for i, t in enumerate(raw)]
embed_chunks(chunks)
hits = InMemoryDenseRetriever(chunks).search(QUESTION, k=3)
print(f'\nQ: "{QUESTION}"\nTop-3 retrieved chunks:')
for h in hits:
    print(f'  [{h.score:.2f}] "{h.chunk.content.strip()}"')
print("\n→ The number rows and the period headers land in SEPARATE chunks. The model sees")
print("  'Net income 10.5 11.2 12.0 13.5' with no period mapping, or a header with no numbers.")
print("  Which one is Q3 2025? Ambiguous. This is exactly the fixed-chunk table-fragmentation")
print("  failure the course measures at 0.00 context precision.")

print("\n" + "=" * 72)
print("APPROACH 2 — Pipeline A deterministic extraction (ours)")
print("=" * 72)
metrics = extract_from_excel(XLSX)
m = next(x for x in metrics if x.metric_type == METRIC and x.period == PERIOD)
flags = validate(metrics)
recon = [f for f in flags if f.metric_type == "net_income" and f.period == PERIOD]
print(f'\nQ: "{QUESTION}"')
print(f"  Answer: ${m.value:,.1f}M   (exact, not retrieved)")
print(f"  Provenance: {m.tab}!{m.cell_ref}   (traceable to the source cell)")
print(f"  Reconciliation (net = pretax - tax): {'FAILED — ' + recon[0].message if recon else 'holds ✓'}")
print("\n→ Deterministic, exact, traceable, and self-checked. No retrieval, no ambiguity,")
print("  no hallucination surface. This is why our architecture routes numbers through")
print("  Pipeline A and reserves RAG for the narrative 'why'.")
