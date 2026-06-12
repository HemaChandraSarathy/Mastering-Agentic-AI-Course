"""Generate SYNTHETIC sample datasets for THREE fictional companies (replaces the Conoco sample).

All figures and text are FABRICATED for demonstration and labeled as such. Each company is written
to corpus/sample/<slug>/ with a messy multi-tab financials .xlsx + narrative .txt docs, plus a
corpus/sample/manifest.json the app reads to ingest each company under its own name.

Companies:
  meridian — Meridian Industrial Components (industrials, rich)
  cobalt   — Cobalt Software (vertical SaaS; ARR/NRR/churn)
  harbor   — Harbor Logistics (logistics; DSO/headcount)
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

REPO = Path(__file__).resolve().parents[1]
SAMPLE = REPO / "corpus" / "sample"
PERIODS = ["Q3 2024", "Q1 2025", "Q2 2025", "Q3 2025"]


def banner(name):
    return f"SAMPLE — {name} (FICTIONAL company, fabricated figures for demo)."


def write_excel(path, name, tabs):
    """tabs = {sheet: (note, [(label,[v0..v3])])}. Messy layout: banner/title/note/blank/header."""
    wb = Workbook()
    first = True
    bold = Font(bold=True)
    for sheet, (note, rows) in tabs.items():
        ws = wb.active if first else wb.create_sheet(sheet)
        ws.title = sheet
        first = False
        ws.append([banner(name)]); ws["A1"].font = Font(bold=True, italic=True)
        ws.append([sheet]); ws["A2"].font = bold
        ws.append([note]); ws.append([])
        ws.append(["Line item"] + PERIODS)
        for c in range(1, len(PERIODS) + 2):
            ws.cell(row=ws.max_row, column=c).font = bold
        for label, vals in rows:
            ws.append([label] + vals)
    wb.save(path)


# ---------------------------------------------------------------------------
# MERIDIAN — industrials (rich narratives)
# ---------------------------------------------------------------------------
MERIDIAN_TABS = {
    "Income Statement": ("$ in millions", [
        ("Revenue", [98.0, 102.0, 105.0, 110.0]),
        ("Total costs and expenses", [84.0, 87.0, 89.0, 92.0]),
        ("Income before income taxes", [14.0, 15.0, 16.0, 18.0]),
        ("Income tax provision", [3.5, 3.8, 4.0, 4.5]),
        ("Net income", [10.5, 11.2, 12.0, 13.5])]),
    "Operating Metrics": ("$M / % / # as noted", [
        ("EBITDA", [18.0, 19.0, 19.8, 20.5]),
        ("Gross margin", [34.0, 34.5, 34.2, 35.0]),
        ("Net working capital", [45.0, 47.0, 48.0, 50.0]),
        ("Headcount", [1180, 1190, 1205, 1220]),
        ("Days sales outstanding", [52, 51, 50, 49])]),
    "Debt & Liquidity": ("$M / x", [
        ("Net debt", [210.0, 205.0, 200.0, 195.0]),
        ("Cash and cash equivalents", [25.0, 28.0, 30.0, 32.0]),
        ("Net leverage", [2.8, 2.7, 2.6, 2.5])]),
}

MERIDIAN_Q3 = """SAMPLE — Meridian Industrial Components, FICTIONAL company. Fabricated for demo.
Q3 2025 Earnings Call Transcript — November 6, 2025

Sarah Chen - Chief Executive Officer
Revenue was $110 million, up 4.8% sequentially and ~12% year over year, led by aerospace and medical
device demand. The headline event was a new five-year supply agreement with a top-three commercial
aerospace OEM for precision-machined titanium structural components, expected to begin contributing in
the second quarter of 2026 and ramp to roughly $40 million of annual revenue by 2028. Gross margin
expanded to 35.0%, helped by first-half pricing actions now fully in effect and easing specialty-alloy,
steel, and freight costs. EBITDA was $20.5 million. We commissioned a second machining cell at our
Dayton, Ohio facility in September, adding ~15% titanium capacity.

David Okafor - Chief Financial Officer
Net income was $13.5 million. Net debt fell to $195 million and net leverage to 2.5x, down from 2.8x at
the start of the year; cash ended at $32 million. Working capital rose to $50 million on titanium
inventory build ahead of the ramp, though days sales outstanding improved to 49. Headcount was 1,220.
Capital priority remains debt paydown until ~2.0x leverage, then bolt-on M&A; no dividend is contemplated.

QUESTIONS AND ANSWERS
Maria Lopez - Crestview Research - Analyst
What is the margin profile of the aerospace program?
Sarah Chen - Chief Executive Officer
Mid-to-high-30s gross margin — above the company average and accretive to mix.
Tom Becker - Harbor Industrial Partners - Analyst
The biggest risk to the 2026 ramp?
Sarah Chen - Chief Executive Officer
OEM qualification and PPAP timing; a one-quarter slip would push out the revenue contribution. We also
expect free cash flow conversion near 50% of EBITDA through the first half of 2026 on the inventory build,
normalizing toward 70% afterward. We are raising full-year revenue guidance to $425–$430 million.
"""

MERIDIAN_Q2 = """SAMPLE — Meridian Industrial Components, FICTIONAL company. Fabricated for demo.
Q2 2025 Earnings Call Transcript — August 7, 2025

Sarah Chen - CEO
Second-quarter revenue was $105 million, up ~3% sequentially. Gross margin was 34.2%, below Q1's 34.5%
on unfavorable mix as lower-margin industrial orders shipped early. EBITDA was $19.8 million.
David Okafor - CFO
Net income $12.0 million; net debt $200 million; net leverage 2.6x; cash $30 million; working capital
$48 million; DSO 50; headcount 1,205. We expect margin to recover in H2 as pricing reads through.
"""

MERIDIAN_ANALYST = """SAMPLE — Meridian Industrial Components, FICTIONAL. Fabricated for demo.
Operating Advisor Update — Q3 2025 (Harbor Industrial Partners)
Rating: Constructive. High-quality quarter: 4.8% sequential growth to $110M, margin to 35.0%,
deleveraging to 2.5x. The five-year aerospace agreement (mid-to-high-30s margin, ~$40M annual run-rate by
2028) lengthens revenue visibility and supports the premium-mix thesis. Risks: 2026 ramp/PPAP timing;
working-capital build holds FCF conversion near 50% of EBITDA through H1 2026; pricing durability if alloy
costs fall. Exit: a credible strategic sale to a tier-one supplier in 2027–2028 at ~2.0x leverage.
"""

# ---------------------------------------------------------------------------
# COBALT — vertical SaaS
# ---------------------------------------------------------------------------
COBALT_TABS = {
    "Income Statement": ("$ in millions", [
        ("Revenue", [40.0, 43.0, 45.0, 48.0]),
        ("Total costs and expenses", [36.0, 38.0, 39.0, 41.0]),
        ("Income before income taxes", [4.0, 5.0, 6.0, 7.0]),
        ("Income tax provision", [1.0, 1.2, 1.5, 1.7]),
        ("Net income", [3.0, 3.8, 4.5, 5.3])]),
    "SaaS Metrics": ("$M / % as noted", [
        ("ARR", [160.0, 172.0, 180.0, 192.0]),
        ("Net revenue retention", [112.0, 113.0, 114.0, 115.0]),
        ("Gross margin", [78.0, 78.5, 79.0, 80.0]),
        ("Customer churn", [8.0, 7.5, 7.0, 6.5]),
        ("Headcount", [320, 330, 345, 360])]),
}
COBALT_Q3 = """SAMPLE — Cobalt Software, FICTIONAL company. Fabricated for demo.
Q3 2025 Earnings Call Transcript

Priya Nair - CEO
ARR reached $192 million, up from $180 million last quarter, with net revenue retention of 115%. Revenue
was $48 million and gross margin improved to 80% on infrastructure efficiency. We closed two seven-figure
enterprise logos in the quarter and reduced logo churn to 6.5%. We are expanding the platform into
payments, which we expect to lift NRR further in 2026.
Daniel Roth - CFO
Net income was $5.3 million. We continue to operate near the Rule-of-40 with ~14% growth and ~15% margin,
and net debt declined to $42 million. Headcount was 360, with most hiring in R&D and customer success.
The main risk we watch is enterprise sales-cycle elongation in a tighter macro.
"""

# ---------------------------------------------------------------------------
# HARBOR — logistics
# ---------------------------------------------------------------------------
HARBOR_TABS = {
    "Income Statement": ("$ in millions", [
        ("Revenue", [220.0, 225.0, 230.0, 240.0]),
        ("Total costs and expenses", [200.0, 205.0, 209.0, 218.0]),
        ("Income before income taxes", [20.0, 20.0, 21.0, 22.0]),
        ("Income tax provision", [5.0, 5.0, 5.2, 5.5]),
        ("Net income", [15.0, 15.0, 15.8, 16.5])]),
    "Operating Metrics": ("$M / % / # as noted", [
        ("EBITDA", [30.0, 31.0, 32.0, 34.0]),
        ("Gross margin", [18.0, 18.2, 18.5, 19.0]),
        ("Net debt", [300.0, 295.0, 290.0, 285.0]),
        ("Cash and cash equivalents", [40.0, 42.0, 44.0, 46.0]),
        ("Days sales outstanding", [45, 44, 43, 42]),
        ("Headcount", [2200, 2220, 2250, 2280])]),
}
HARBOR_Q3 = """SAMPLE — Harbor Logistics, FICTIONAL company. Fabricated for demo.
Q3 2025 Earnings Call Transcript

Marcus Hale - CEO
Revenue was $240 million, up ~4% sequentially as freight volumes stabilized and our dedicated-fleet
contracts expanded. EBITDA reached $34 million at an 14.2% margin. We signed a multi-year dedicated
contract with a national retailer that adds committed volume beginning in 2026.
Elena Vargas - CFO
Net income was $16.5 million. Net debt declined to $285 million as we paid down the revolver; cash was
$46 million. Days sales outstanding improved to 42 on tighter collections. Headcount was 2,280. The main
risks are diesel price volatility and driver availability, the key operational constraint in the sector.
"""

COBALT_Q2 = """SAMPLE — Cobalt Software, FICTIONAL company. Fabricated for demo.
Q2 2025 Earnings Call Transcript

Priya Nair - CEO
ARR was $180 million, up from $172 million, with net revenue retention of 114%. Revenue was $45 million and
gross margin 79%. We launched our analytics module this quarter and saw early attach in the mid-market.
Daniel Roth - CFO
Net income was $4.5 million. Logo churn was 7.0%. Net debt was $45 million and headcount 345. We reiterate
our full-year target of approximately $200 million ARR exiting 2025.
"""

COBALT_ANALYST = """SAMPLE — Cobalt Software, FICTIONAL. Fabricated for demo.
Sponsor Update — Q3 2025
Rating: On plan. Cobalt is running a clean SaaS playbook: ARR to $192M (+7% QoQ), net revenue retention 115%,
gross margin 80%, logo churn down to 6.5%. The payments expansion is the key 2026 upsell lever and supports
the NRR trajectory. Risks: enterprise sales-cycle elongation in a tighter macro, and concentration in the two
large new logos. Exit: a scaled vertical-SaaS asset above 115% NRR is a credible strategic or
sponsor-to-sponsor sale in 2027.
"""

HARBOR_Q2 = """SAMPLE — Harbor Logistics, FICTIONAL company. Fabricated for demo.
Q2 2025 Earnings Call Transcript

Marcus Hale - CEO
Revenue was $230 million, up modestly as spot freight stayed soft while dedicated-fleet contracts grew.
EBITDA was $32 million. We added two dedicated lanes with existing customers.
Elena Vargas - CFO
Net income $15.8 million. Net debt $290 million; cash $44 million; days sales outstanding 43; headcount 2,250.
We expect margin to improve in the second half as the dedicated mix increases and diesel costs stabilize.
"""

HARBOR_ANALYST = """SAMPLE — Harbor Logistics, FICTIONAL. Fabricated for demo.
Operating Partner Update — Q3 2025
Rating: Constructive. Harbor is shifting mix toward dedicated contract logistics, which lifted EBITDA margin
to ~14%. The new national-retailer dedicated contract (committed volume from 2026) de-risks the revenue base.
Risks: diesel price volatility and driver availability — the binding operational constraint in trucking. Net
debt is trending down ($300M → $285M over three quarters). Exit: a strategic sale to a larger 3PL on improved
margins and contracted backlog.
"""

COMPANIES = [
    {"slug": "meridian", "name": "Meridian Industrial Components",
     "sector": "Industrials — Precision Components", "tabs": MERIDIAN_TABS,
     "narratives": {
         "Meridian_Q3_2025_Earnings_Call_Transcript.txt": MERIDIAN_Q3,
         "Meridian_Q2_2025_Earnings_Call_Transcript.txt": MERIDIAN_Q2,
         "Meridian_Q3_2025_Analyst_Report.txt": MERIDIAN_ANALYST}},
    {"slug": "cobalt", "name": "Cobalt Software", "sector": "Software — Vertical SaaS",
     "tabs": COBALT_TABS,
     "narratives": {"Cobalt_Q3_2025_Earnings_Call_Transcript.txt": COBALT_Q3,
                    "Cobalt_Q2_2025_Earnings_Call_Transcript.txt": COBALT_Q2,
                    "Cobalt_Q3_2025_Analyst_Report.txt": COBALT_ANALYST}},
    {"slug": "harbor", "name": "Harbor Logistics", "sector": "Logistics — Trucking",
     "tabs": HARBOR_TABS,
     "narratives": {"Harbor_Q3_2025_Earnings_Call_Transcript.txt": HARBOR_Q3,
                    "Harbor_Q2_2025_Earnings_Call_Transcript.txt": HARBOR_Q2,
                    "Harbor_Q3_2025_Analyst_Report.txt": HARBOR_ANALYST}},
]


def main():
    manifest = []
    for co in COMPANIES:
        d = SAMPLE / co["slug"]
        d.mkdir(parents=True, exist_ok=True)
        write_excel(d / f"{co['slug']}_SAMPLE_financials.xlsx", co["name"], co["tabs"])
        for fname, text in co["narratives"].items():
            (d / fname).write_text(text, encoding="utf-8")
        manifest.append({"slug": co["slug"], "name": co["name"], "sector": co["sector"],
                         "dir": f"corpus/sample/{co['slug']}"})
    (SAMPLE / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(f"Wrote {len(COMPANIES)} sample companies to {SAMPLE.relative_to(REPO)}:")
    for co in COMPANIES:
        print(f"  {co['name']} → corpus/sample/{co['slug']}/")


if __name__ == "__main__":
    main()
