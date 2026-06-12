"""Build a deliberately MESSY multi-tab Excel from REAL ConocoPhillips figures.

WHY: the Conoco corpus PDFs have no text layer (they render as images), and messy Excel
is half the real PE-ingestion pain. This script re-keys ConocoPhillips' *real, public*
income-statement figures into a fabricated, inconsistent multi-tab workbook so Pipeline A
has a clean deterministic input AND the demo looks like a real portco drop.

INTEGRITY:
  - The NUMBERS are real, taken from ConocoPhillips' quarterly Consolidated Income
    Statements (10-Q, Item 1). Source = the Conoco Sample Data / Quarterly Reports PDFs.
  - The LAYOUT (tabs, headers, ordering, label spellings) is FABRICATED to be messy.
  - The file is labeled SYNTHETIC on every tab. It is a placeholder for a private portco;
    it is not real customer data.

Figures captured (3-month periods, $ in millions unless noted):
                          Q3 2024   Q1 2025   Q2 2025   Q3 2025
  Sales & other op rev     13,041    16,517    14,004    15,031
  Total revenues & income  13,604    17,101    14,740    15,522
  Total costs & expenses   10,369    12,635    11,723    12,594
  Income before tax         3,235     4,466     3,017     2,928
  Income tax provision      1,176     1,617     1,046     1,202
  Net income                2,059     2,849     1,971     1,726
  Diluted EPS ($/sh)         1.76      2.32      1.56      1.38
  -- cost detail (unmapped on purpose) --
  Purchased commodities     4,747     6,188     5,085     5,857
  Production & operating    2,261     2,506     2,572     2,632
  SG&A                        186       191       250       271
  DD&A                      2,390     2,746     2,838     2,917
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

REPO_ROOT = Path(__file__).resolve().parents[1]
OUT = REPO_ROOT / "corpus" / "ConocoPhillips_SYNTHETIC_financials.xlsx"

PERIODS = ["Q3 2024", "Q1 2025", "Q2 2025", "Q3 2025"]

INCOME_STATEMENT = [
    ("Sales and other operating revenues", [13041, 16517, 14004, 15031]),
    ("Total revenues and other income", [13604, 17101, 14740, 15522]),
    ("Total costs and expenses", [10369, 12635, 11723, 12594]),
    ("Income before income taxes", [3235, 4466, 3017, 2928]),
    ("Income tax provision", [1176, 1617, 1046, 1202]),
    ("Net income", [2059, 2849, 1971, 1726]),
]

PER_SHARE = [
    ("Diluted EPS", [1.76, 2.32, 1.56, 1.38]),
]

COST_DETAIL = [
    ("Purchased commodities", [4747, 6188, 5085, 5857]),
    ("Production and operating expenses", [2261, 2506, 2572, 2632]),
    ("Selling, general and administrative expenses", [186, 191, 250, 271]),
    ("Depreciation, depletion and amortization", [2390, 2746, 2838, 2917]),
]

BANNER = "SYNTHETIC PLACEHOLDER — real ConocoPhillips public figures, fabricated messy layout. Not customer data."


def _write_tab(ws, title, note, period_headers, rows, *, label_header):
    """Write a deliberately messy tab: banner row, note row, blank row, header, data."""
    bold = Font(bold=True)
    ws.append([BANNER])
    ws["A1"].font = Font(bold=True, italic=True)
    ws.append([title])
    ws["A2"].font = bold
    ws.append([note])
    ws.append([])  # blank separator row — extractor must skip it
    header = [label_header] + period_headers
    ws.append(header)
    for c in range(1, len(header) + 1):
        ws.cell(row=ws.max_row, column=c).font = bold
    for label, values in rows:
        ws.append([label] + values)


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Tab 1 — Income Statement (clean-ish period headers)
    ws1 = wb.active
    ws1.title = "Income Statement"
    _write_tab(
        ws1,
        "Consolidated Income Statement",
        "$ in millions",
        PERIODS,
        INCOME_STATEMENT,
        label_header="Line item",
    )

    # Tab 2 — Per Share (different header style: dashes, different label col name)
    ws2 = wb.create_sheet("Per Share Data")
    _write_tab(
        ws2,
        "Per-share data",
        "USD per diluted share",
        [p.replace(" ", "-") for p in PERIODS],   # "Q3-2024" — inconsistent on purpose
        PER_SHARE,
        label_header="Metric",
    )

    # Tab 3 — Cost Detail (lines that intentionally DON'T map to the taxonomy)
    ws3 = wb.create_sheet("Cost Detail")
    _write_tab(
        ws3,
        "Costs and expenses detail",
        "$ in millions",
        PERIODS,
        COST_DETAIL,
        label_header="Expense line",
    )

    wb.save(OUT)
    print(f"Wrote {OUT.relative_to(REPO_ROOT)}  ({len(wb.sheetnames)} tabs: {wb.sheetnames})")


if __name__ == "__main__":
    main()
