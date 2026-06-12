"""Deterministic extraction from grids (Excel sheets, PDF tables). NO LLM, NO embeddings.

Strategy for messy grids:
  - Find the HEADER row = the first row containing >=2 parseable period tokens (via
    parse_period). Columns under those tokens are period columns.
  - The LABEL column is the first non-empty column at/left of the first period column.
  - Each subsequent row with a non-empty label + numeric cells becomes one ExtractedMetric
    per period, mapped through the taxonomy. Unmapped labels are kept (metric_type=None).

`metrics_from_grid` is the shared engine; `extract_from_excel` (openpyxl) and
`extract_from_text_pdf` (pdfplumber tables) both feed it a 2-D grid + a provenance ref fn.
This handles banner/title/note/blank rows above the header without hardcoding offsets.
"""
from __future__ import annotations

from pathlib import Path
from typing import Callable, Optional, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..taxonomy import map_label, ALL_METRICS
from .models import ExtractedMetric, parse_period

Grid = Sequence[Sequence[object]]   # rows x cols of cell values (None for empty)


def _to_number(cell_value) -> float | None:
    if cell_value is None:
        return None
    if isinstance(cell_value, (int, float)):
        return float(cell_value)
    s = str(cell_value).strip().replace(",", "").replace("$", "")
    if s in ("", "—", "-", "–", "n/a", "na"):
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if neg else v


def metrics_from_grid(
    grid: Grid,
    *,
    source_file: str,
    tab: Optional[str],
    ref_fn: Callable[[int, int], str],
    confidence: Optional[float] = None,
) -> list[ExtractedMetric]:
    """Shared grid engine: 2-D cell grid -> ExtractedMetric rows.

    `grid[r][c]` is a raw cell value (None for empty). `ref_fn(r, c)` returns a provenance
    string for that cell (e.g. 'B7' for Excel, 'p2:t1[3,4]' for a PDF table). Used by both
    the Excel and PDF-table extractors so the header/label/period logic lives in one place.
    """
    def _cell(row: Sequence[object], c: int) -> object:
        return row[c] if c < len(row) else None

    # 1) locate header row = first row with >=2 parseable period tokens
    header_idx: Optional[int] = None
    period_cols: dict[int, str] = {}     # 0-based column index -> canonical period
    for r_idx, row in enumerate(grid):
        found: dict[int, str] = {}
        for c_idx in range(len(row)):
            v = _cell(row, c_idx)
            p = parse_period(v) if v is not None else None
            if p:
                found[c_idx] = p
        if len(found) >= 2:
            header_idx, period_cols = r_idx, found
            break
    if header_idx is None:
        return []

    first_period_col = min(period_cols)
    # 2) label column = first non-empty column strictly left of the first period col
    label_col = 0
    for c in range(first_period_col):
        if any(_cell(grid[r], c) not in (None, "") for r in range(header_idx + 1, len(grid))):
            label_col = c
            break

    # 3) data rows
    out: list[ExtractedMetric] = []
    for r_idx in range(header_idx + 1, len(grid)):
        row = grid[r_idx]
        lv = _cell(row, label_col)
        label = str(lv).strip() if lv is not None else ""
        if not label:
            continue
        metric_type = map_label(label)
        unit = ALL_METRICS[metric_type].unit.value if metric_type else None
        for c_idx, period in period_cols.items():
            num = _to_number(_cell(row, c_idx))
            if num is None:
                continue
            out.append(ExtractedMetric(
                raw_label=label, metric_type=metric_type, period=period, value=num,
                unit=unit, source_file=source_file, tab=tab,
                cell_ref=ref_fn(r_idx, c_idx), confidence=confidence,
            ))
    return out


def extract_from_excel(path: str | Path) -> list[ExtractedMetric]:
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    out: list[ExtractedMetric] = []
    for ws in wb.worksheets:
        grid = [[cell.value for cell in row] for row in ws.iter_rows()]
        out += metrics_from_grid(
            grid, source_file=path.name, tab=ws.title,
            ref_fn=lambda r, c: f"{get_column_letter(c + 1)}{r + 1}",
        )
    return out


def extract_from_text_pdf(path: str | Path) -> list[ExtractedMetric]:
    """Deterministic table extraction from a TEXT-layer PDF (pdfplumber). NO OCR, NO LLM.

    For text PDFs that contain financial tables (vs the image-only path which needs OCR).
    Each detected table is a grid fed to `metrics_from_grid`; non-financial tables (no period
    header, e.g. a transcript's layout) simply yield nothing. Numbers from a text PDF go
    through Pipeline A — never via narrative RAG.
    """
    import pdfplumber

    path = Path(path)
    out: list[ExtractedMetric] = []
    with pdfplumber.open(path) as pdf:
        for pno, page in enumerate(pdf.pages, start=1):
            for tno, table in enumerate(page.extract_tables() or []):
                out += metrics_from_grid(
                    table, source_file=path.name, tab=f"page {pno} table {tno + 1}",
                    ref_fn=lambda r, c, _p=pno, _t=tno: f"p{_p}:t{_t + 1}[{r},{c}]",
                )
    return out
